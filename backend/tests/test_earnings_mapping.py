import unittest
from io import BytesIO

from openpyxl import load_workbook

from app.models.schemas import TaskLogRow, coerce_task_log
from app.services.analytics_service import build_analytics_summary
from app.services.earnings import compute_hours
from app.services.export_service import build_csv, build_xlsx, normalize_export_rows


class EarningsMappingTests(unittest.TestCase):
    def test_zero_usd_backfill_uses_positive_canonical_earnings(self) -> None:
        raw = {
            "calculated_earnings_usd": 0,
            "calculated_earnings": 12.5,
        }

        coerced = coerce_task_log(raw)
        exported = normalize_export_rows([raw])[0]
        summary = build_analytics_summary([raw])

        self.assertEqual(coerced["earnings_usd"], 12.5)
        self.assertEqual(exported["earnings_usd"], 12.5)
        self.assertEqual(summary.kpis.total_earned, 12.5)

    def test_legitimate_zero_earnings_stays_zero(self) -> None:
        raw = {
            "calculated_earnings_usd": 0,
            "calculated_earnings": 0,
            "payment_status": "paid",
        }

        self.assertEqual(coerce_task_log(raw)["earnings_usd"], 0)
        self.assertEqual(normalize_export_rows([raw])[0]["earnings_usd"], 0)
        summary = build_analytics_summary([raw])
        self.assertEqual(summary.kpis.total_earned, 0)
        self.assertEqual(summary.kpis.total_paid, 0)
        self.assertEqual(summary.kpis.total_pending, 0)

    def test_all_pending_earnings_are_pending(self) -> None:
        summary = build_analytics_summary(
            [
                {"calculated_earnings_usd": 5},
                {"calculated_earnings_usd": 7.5, "payment_status": "pending"},
            ]
        )

        self.assertEqual(summary.kpis.total_earned, 12.5)
        self.assertEqual(summary.kpis.total_paid, 0)
        self.assertEqual(summary.kpis.total_pending, 12.5)
        self.assertEqual(summary.series[0].paid, 0)
        self.assertEqual(summary.series[0].pending, 12.5)

    def test_all_paid_earnings_are_paid(self) -> None:
        summary = build_analytics_summary(
            [{"calculated_earnings_usd": 12.5, "payment_status": "paid"}]
        )

        self.assertEqual(summary.kpis.total_earned, 12.5)
        self.assertEqual(summary.kpis.total_paid, 12.5)
        self.assertEqual(summary.kpis.total_pending, 0)
        self.assertEqual(summary.series[0].paid, 12.5)
        self.assertEqual(summary.series[0].pending, 0)

    def test_mixed_payment_statuses_split_gross_earnings(self) -> None:
        summary = build_analytics_summary(
            [
                {"calculated_earnings_usd": 10, "payment_status": "paid"},
                {"calculated_earnings_usd": 4.25, "payment_status": "pending"},
            ]
        )

        self.assertEqual(summary.kpis.total_earned, 14.25)
        self.assertEqual(summary.kpis.total_paid, 10)
        self.assertEqual(summary.kpis.total_pending, 4.25)
        self.assertEqual(summary.series[0].earnings, 14.25)
        self.assertEqual(summary.series[0].paid, 10)
        self.assertEqual(summary.series[0].pending, 4.25)

    def test_missing_or_invalid_status_defaults_to_pending(self) -> None:
        self.assertEqual(coerce_task_log({})["payment_status"], "pending")
        self.assertIsNone(coerce_task_log({})["paid_at"])
        coerced = coerce_task_log(
            {"payment_status": "not-a-status", "paid_at": "not-a-timestamp"}
        )
        self.assertEqual(coerced["payment_status"], "pending")
        self.assertIsNone(coerced["paid_at"])
        model = TaskLogRow(payment_status=None, paid_at="not-a-timestamp")
        self.assertEqual(model.payment_status.value, "pending")
        self.assertIsNone(model.paid_at)

    def test_payment_fields_are_in_csv_and_xlsx_exports(self) -> None:
        rows = [
            {
                "calculated_earnings_usd": 12.5,
                "payment_status": "paid",
                "paid_at": "2026-08-07T12:30:00Z",
            }
        ]

        csv_text = build_csv(rows).decode("utf-8-sig")
        self.assertIn("payment_status", csv_text)
        self.assertIn("paid_at", csv_text)
        self.assertIn("earnings_usd", csv_text)
        self.assertIn("currency_code", csv_text)
        self.assertIn("12.5", csv_text)

        workbook = load_workbook(BytesIO(build_xlsx(rows)), read_only=True)
        headers = list(next(workbook.active.iter_rows(values_only=True)))
        self.assertIn("payment_status", headers)
        self.assertIn("paid_at", headers)
        self.assertIn("earnings_usd", headers)
        self.assertIn("currency_code", headers)

        workbook.close()

    def test_positive_usd_snapshot_remains_authoritative(self) -> None:
        raw = {
            "calculated_earnings_usd": 7.5,
            "calculated_earnings": 12.5,
        }

        self.assertEqual(coerce_task_log(raw)["earnings_usd"], 7.5)
        self.assertEqual(normalize_export_rows([raw])[0]["earnings_usd"], 7.5)

    def test_minutes_formula_uses_sixty_minutes_per_hour(self) -> None:
        self.assertEqual(compute_hours(2, 4, 30, 15), 2.0)


if __name__ == "__main__":
    unittest.main()
